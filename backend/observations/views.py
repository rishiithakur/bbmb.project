from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
from .models import WaterLevelLog, DamObservation, ObservationPhoto
from .serializers import WaterLevelLogSerializer, DamObservationSerializer
from accounts.permissions import IsBBMCAdmin
from audit.utils import log_audit


class WaterLevelLogViewSet(viewsets.ModelViewSet):
    """
    Main ViewSet for water level observations.

    Key behaviours:
    - Operators are scoped to their assigned site automatically.
    - DRAFT saves allow partial data (no water_level_m required).
    - FINAL submissions require water_level_m and lock the record.
    - Duplicate check only for FINAL on same site+date.
    - /my-submissions/ returns only the current operator's records.
    - /export-excel/ is admin-only and returns .xlsx.
    """
    serializer_class = WaterLevelLogSerializer
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'log_id'

    # ── Queryset ──────────────────────────────────────────────────────────────
    def get_queryset(self):
        user = self.request.user
        qs = WaterLevelLog.objects.select_related('site', 'user', 'dam_detail').order_by('-observation_datetime')

        if user.role in ['admin', 'supreme_admin', 'ultra_admin', 'viewer']:
            # Apply optional filters from query params
            return self._apply_filters(qs)

        # Operator: scope to their site
        if user.assigned_site_id:
            qs = qs.filter(site_id=user.assigned_site_id)
            return self._apply_filters(qs)

        return WaterLevelLog.objects.none()

    def _apply_filters(self, qs):
        params = self.request.query_params
        if 'site' in params:
            qs = qs.filter(site_id=params['site'])
        if 'operator' in params:
            qs = qs.filter(user_id=params['operator'])
        if 'status' in params:
            status_val = params['status'].upper()
            if status_val == 'FINAL':
                qs = qs.filter(is_verified=True)
            elif status_val == 'DRAFT':
                qs = qs.filter(is_verified=False)
        if 'date_from' in params:
            qs = qs.filter(observation_date__gte=params['date_from'])
        if 'date_to' in params:
            qs = qs.filter(observation_date__lte=params['date_to'])
        if 'alert_level' in params:
            qs = qs.filter(dam_detail__alert_level=params['alert_level'])
        if params.get('mine') == 'true':
            qs = qs.filter(user=self.request.user)
        return qs

    # ── Create ────────────────────────────────────────────────────────────────
    def create(self, request, *args, **kwargs):
        user = request.user

        # Assign site from user profile for operators
        if user.role not in ['admin', 'supreme_admin', 'ultra_admin']:
            if not user.assigned_site_id:
                return Response(
                    {'error': 'You are not assigned to any monitoring site.'},
                    status=status.HTTP_403_FORBIDDEN
                )
            # Inject site into mutable copy
            data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)
            data['site'] = user.assigned_site_id
            request._full_data = data

        # Duplicate check only for FINAL submissions
        submitted_status = request.data.get('status', 'DRAFT').upper()
        if submitted_status == 'FINAL':
            site = request.data.get('site') or getattr(user, 'assigned_site_id', None)
            obs_date = request.data.get('observation_date')
            if site and obs_date:
                if WaterLevelLog.objects.filter(site=site, observation_date=obs_date, is_verified=True).exists():
                    return Response({
                        'warning': 'A finalized observation already exists for this site and date.',
                        'duplicate': True
                    }, status=status.HTTP_409_CONFLICT)

        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        log = serializer.save(user=self.request.user)
        action_type = 'ENTRY_FINALIZED' if log.is_verified else 'ENTRY_DRAFT_CREATED'
        self._log_audit(action_type, log, serializer.data)
        self._trigger_submission_notifications(log, is_new=True)

    def _trigger_submission_notifications(self, log, is_new=False):
        try:
            from audit.models import Notification
            site = log.site
            user = log.user
            
            # 1. Notify the submitting operator
            if log.is_verified:
                Notification.objects.create(
                    recipient_user=user,
                    site=site,
                    type='FINAL_SUBMISSION',
                    title='Submission Finalized',
                    message=f"Your telemetry submission for site '{site.station_name}' on {log.observation_date} has been successfully finalized and locked.",
                    metadata_json={'log_id': log.log_id, 'water_level': float(log.water_level_m or 0)}
                )

                # 2. Notify all administrators
                Notification.objects.create(
                    recipient_role='admin',
                    site=site,
                    type='FINAL_SUBMISSION_ALERT',
                    title='Telemetry Submission: Finalized',
                    message=f"Operator '{user.username}' has submitted final telemetry data for site '{site.station_name}'. Current level is {log.water_level_m}m.",
                    metadata_json={'log_id': log.log_id, 'operator': user.username, 'water_level': float(log.water_level_m or 0)}
                )
            else:
                Notification.objects.create(
                    recipient_user=user,
                    site=site,
                    type='DRAFT_SAVED',
                    title='Draft Telemetry Saved',
                    message=f"Draft telemetry data for site '{site.station_name}' has been successfully autosaved to the database.",
                    metadata_json={'log_id': log.log_id}
                )

            # 3. Check for reservoir warnings/alerts if there is a dam_detail
            dam = getattr(log, 'dam_detail', None)
            if dam and dam.alert_level in ['Yellow', 'Orange', 'Red']:
                alert_type = 'RESERVOIR_ALERT'
                title = f"Critical Alert: {dam.alert_level} Level Detected"
                message = f"Warning: Station '{site.station_name}' is at {dam.alert_level} alert level! Current level is {log.water_level_m}m. Condition is reported as: '{dam.dam_condition}'."
                
                # Notify operator
                Notification.objects.create(
                    recipient_user=user,
                    site=site,
                    type=alert_type,
                    title=title,
                    message=message,
                    metadata_json={'log_id': log.log_id, 'alert_level': dam.alert_level, 'water_level': float(log.water_level_m or 0)}
                )

                # Notify admins
                Notification.objects.create(
                    recipient_role='admin',
                    site=site,
                    type=alert_type,
                    title=title,
                    message=message,
                    metadata_json={'log_id': log.log_id, 'alert_level': dam.alert_level, 'operator': user.username}
                )
        except Exception as e:
            print(f"Error in _trigger_submission_notifications: {e}")

    # ── Update ────────────────────────────────────────────────────────────────
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        user = request.user

        # Guard: finalized records locked for operators
        if instance.is_verified and user.role not in ['admin', 'supreme_admin', 'ultra_admin']:
            return Response(
                {'error': 'Finalized records cannot be edited.'},
                status=status.HTTP_403_FORBIDDEN
            )
        # Guard: site isolation
        if user.role not in ['admin', 'supreme_admin', 'ultra_admin']:
            if instance.site_id != user.assigned_site_id:
                return Response(
                    {'error': 'You cannot edit records from other sites.'},
                    status=status.HTTP_403_FORBIDDEN
                )

        return super().update(request, *args, **kwargs)

    def perform_update(self, serializer):
        old_verified = self.get_object().is_verified
        log = serializer.save()
        if log.is_verified and not old_verified:
            action_type = 'ENTRY_FINALIZED'
        elif not log.is_verified:
            action_type = 'ENTRY_DRAFT_EDITED'
        else:
            action_type = 'ENTRY_VERIFIED'
        self._log_audit(action_type, log, serializer.data)
        self._trigger_submission_notifications(log, is_new=False)

    # ── Custom Actions ────────────────────────────────────────────────────────

    @action(detail=False, methods=['get'], url_path='my-submissions')
    def my_submissions(self, request):
        """Returns the current operator's own observation history."""
        qs = WaterLevelLog.objects.filter(
            user=request.user
        ).select_related('site', 'dam_detail').order_by('-observation_datetime')

        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='stats')
    def stats(self, request):
        from observations.models import VwDashboardSummary
        summary = VwDashboardSummary.objects.first()

        if not summary:
            return Response({
                'total_sites': 0, 'total_users': 0, 'last_24h_entries': 0,
                'red_alerts': 0, 'yellow_alerts': 0, 'normal_sites': 0,
                'health_score': 100, 'system_status': 'Online',
                'pending_verifications': 0,
            })

        health_score = max(0, 100 - int(summary.red_alert_sites or 0) * 15)
        return Response({
            'total_sites': summary.total_sites,
            'total_users': summary.total_operators,
            'last_24h_entries': summary.total_entries,
            'red_alerts': summary.red_alert_sites,
            'yellow_alerts': 0,
            'pending_verifications': summary.pending_verifications,
            'normal_sites': max(0, (summary.total_sites or 0) - (summary.red_alert_sites or 0)),
            'health_score': health_score,
            'system_status': 'Online',
        })

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Admin-only Excel export of all filtered observations."""
        if not IsBBMCAdmin().has_permission(request, self):
            return Response({'detail': 'Admin access required.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
        except ImportError:
            return Response(
                {'detail': 'openpyxl is not installed. Run: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Dam Observations'

        # Header styling
        header_fill = PatternFill(start_color='1C2833', end_color='1C2833', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        headers = [
            'Log ID', 'Site Code', 'Station Name', 'Operator', 'Date', 'Time',
            'Water Level (m)', 'Water Level (ft)', 'Storage (MCM)', 'Storage (%)',
            'Inflow (cusecs)', 'Outflow (cusecs)',
            'Spillway Gates Open', 'Spillway Discharge (cusecs)',
            'Power Generation (MW)', 'Rainfall Today (mm)',
            'Weather Condition', 'Alert Level', 'Dam Condition',
            'Remarks', 'Notes', 'Status', 'Verified At',
        ]

        ws.append(headers)
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Data rows — use filtered queryset
        qs = self._apply_filters(
            WaterLevelLog.objects.select_related('site', 'user', 'dam_detail').all()
        ).order_by('-observation_date')

        for log in qs:
            dam = getattr(log, 'dam_detail', None)
            ws.append([
                log.log_id,
                log.site.site_code if log.site else '',
                log.site.station_name if log.site else '',
                log.user.full_name or log.user.username if log.user else '',
                str(log.observation_date) if log.observation_date else '',
                str(log.observation_time) if log.observation_time else '',
                float(log.water_level_m) if log.water_level_m is not None else '',
                log.water_level_ft or '',
                float(log.storage_mcm) if log.storage_mcm is not None else '',
                float(log.storage_percent) if log.storage_percent is not None else '',
                float(log.inflow_cusecs) if log.inflow_cusecs is not None else '',
                float(log.outflow_cusecs) if log.outflow_cusecs is not None else '',
                dam.spillway_gates_open if dam else '',
                float(dam.spillway_discharge_cusecs) if dam and dam.spillway_discharge_cusecs else '',
                float(dam.power_generation_mw) if dam and dam.power_generation_mw else '',
                float(dam.rainfall_today_mm) if dam and dam.rainfall_today_mm else '',
                log.weather_condition or '',
                dam.alert_level if dam else '',
                dam.dam_condition if dam else '',
                log.remarks or '',
                dam.notes if dam else '',
                'FINAL' if log.is_verified else 'DRAFT',
                str(log.verified_at) if log.verified_at else '',
            ])

        # Auto-fit columns
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
        ws.freeze_panes = 'A2'

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        log_audit(request=request, action='REPORT_EXPORT_EXCEL', status='success')

        try:
            from audit.models import Notification
            Notification.objects.create(
                recipient_role='admin',
                type='EXPORT_SUCCESS',
                title='Excel Report Exported',
                message=f"Administrator '{request.user.username}' has successfully exported a consolidated Excel spreadsheet containing {qs.count()} telemetry records."
            )
        except Exception:
            pass

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="dam_observations.xlsx"'
        return response

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _log_audit(self, action_type, log, data=None):
        try:
            log_audit(
                request=self.request,
                action=action_type,
                table_name='water_level_log',
                record_id=log.log_id,
                new_data=data,
                site=log.site
            )
        except Exception:
            pass  # Never let audit failures break the main workflow


class DamObservationViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = DamObservation.objects.all()
    serializer_class = DamObservationSerializer
    permission_classes = [permissions.IsAuthenticated]
