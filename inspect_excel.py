import openpyxl
import json

def inspect_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    structure = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(cell.value)
        
        sample_data = []
        for row in sheet.iter_rows(min_row=2, max_row=4, values_only=True):
            if any(row):
                sample_data.append(list(row))
        
        structure[sheet_name] = {
            "headers": headers,
            "sample_data": sample_data
        }
    return structure

if __name__ == "__main__":
    file_path = r"d:\BBMC DAM WATER LEVEL MONITORING SYSTEM\Dam_WaterLevel_DB_Design.xlsx"
    try:
        data = inspect_excel(file_path)
        with open("excel_structure.json", "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        print("Success: excel_structure.json created.")
    except Exception as e:
        print(f"Error: {e}")
