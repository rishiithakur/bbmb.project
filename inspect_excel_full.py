import openpyxl
import json

def inspect_excel_full(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    structure = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            if any(row):
                data.append(list(row))
        structure[sheet_name] = data
    return structure

if __name__ == "__main__":
    file_path = r"d:\BBMC DAM WATER LEVEL MONITORING SYSTEM\Dam_WaterLevel_DB_Design.xlsx"
    try:
        data = inspect_excel_full(file_path)
        with open("excel_full_data.json", "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        print("Success: excel_full_data.json created.")
    except Exception as e:
        print(f"Error: {e}")
